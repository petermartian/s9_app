import streamlit as st
import pandas as pd
from datetime import date
import plotly.express as px
from fpdf import FPDF
import io
import xlsxwriter
from openpyxl import load_workbook

# --- Excel File Setup ---
EXCEL_FILE = "Daily Transaction.xlsx"
TAB_NAME = "expenses"

# --- Cache Excel File Loading for Performance ---
@st.cache_data
def load_excel_file(file_path, sheet_name):
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except (FileNotFoundError, ValueError) as e:
        st.error(f"❌ Failed to load {sheet_name}: {e}")
        return pd.DataFrame()

# --- Initialize Session State for Form Inputs ---
def initialize_form_state():
    defaults = {
        "expense_date": date.today(),
        "expense_description": "",
        "amount_ngn": 0.0,
        "amount_usd": 0.0
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# Initialize session state
initialize_form_state()

# --- Form UI ---
def run_page():
    st.subheader("🧾 Expense Entry")

    with st.form("expense_form", clear_on_submit=False):
        col1, col2 = st.columns(2)
        with col1:
            expense_date = st.date_input("Date", value=st.session_state.expense_date, key="expense_date")
            description = st.text_input("Expense Description", value=st.session_state.expense_description, key="expense_description")
        with col2:
            amount_ngn = st.number_input("Amount NGN", min_value=0.0, value=st.session_state.amount_ngn, key="amount_ngn")
            amount_usd = st.number_input("Amount USD", min_value=0.0, value=st.session_state.amount_usd, key="amount_usd")

        submitted = st.form_submit_button("✅ Submit Expense")

    # --- Form Submission Logic ---
    if submitted:
        # Input validation
        if st.session_state.amount_ngn < 0 or st.session_state.amount_usd < 0:
            st.error("❌ Amount NGN and Amount USD must be non-negative.")
        else:
            try:
                new_row = {
                    "Date": st.session_state.expense_date.strftime("%Y-%m-%d"),
                    "Expense Description": st.session_state.expense_description,
                    "Amount NGN": round(st.session_state.amount_ngn, 2),
                    "Amount USD": round(st.session_state.amount_usd, 2)
                }

                # Load existing data or create empty DataFrame
                try:
                    book = load_workbook(EXCEL_FILE)
                    if TAB_NAME in book.sheetnames:
                        df_existing = pd.read_excel(EXCEL_FILE, sheet_name=TAB_NAME)
                    else:
                        df_existing = pd.DataFrame(columns=new_row.keys())
                except (FileNotFoundError, ValueError):
                    df_existing = pd.DataFrame(columns=new_row.keys())

                # Append new row
                df_combined = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)

                # Write to Excel without overwriting other sheets
                try:
                    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        df_combined.to_excel(writer, sheet_name=TAB_NAME, index=False)
                    st.success("✅ Expense submitted successfully!")
                except PermissionError:
                    st.error("❌ Cannot write to Excel file: File is open or lacks write permissions.")
                except Exception as e:
                    st.error(f"❌ Failed to write to Excel: {e}")

            except Exception as e:
                st.error(f"❌ Error processing submission: {e}")

    # --- Load Data ---
    df = load_excel_file(EXCEL_FILE, TAB_NAME)

    # Normalize headers
    if not df.empty:
        df.columns = [col.strip().title() for col in df.columns]

        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df[df["Date"].notna()]
        if not df.empty and df["Date"].notna().any():
            start = st.date_input("📅 Start Date", df["Date"].min().date(), key="start_expense")
            end = st.date_input("📅 End Date", df["Date"].max().date(), key="end_expense")
        else:
            start = st.date_input("📅 Start Date", date.today(), key="start_expense")
            end = st.date_input("📅 End Date", date.today(), key="end_expense")
        
        df_filtered = df[(df["Date"] >= pd.to_datetime(start)) & (df["Date"] <= pd.to_datetime(end))]

        for col in ["Amount Ngn", "Amount Usd"]:
            df_filtered[col] = pd.to_numeric(df_filtered[col], errors="coerce")

        st.markdown("### 📊 Expense Summary")
        df_filtered["Week"] = df_filtered["Date"].dt.isocalendar().week
        df_filtered["Month"] = df_filtered["Date"].dt.month

        summary = {
            "Period": ["Daily", "Weekly", "Monthly"],
            "Amount NGN": [
                df_filtered[df_filtered["Date"] == pd.to_datetime(date.today())]["Amount Ngn"].sum(),
                df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["Amount Ngn"].sum(),
                df_filtered[df_filtered["Month"] == date.today().month]["Amount Ngn"].sum()
            ],
            "Amount USD": [
                df_filtered[df_filtered["Date"] == pd.to_datetime(date.today())]["Amount Usd"].sum(),
                df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["Amount Usd"].sum(),
                df_filtered[df_filtered["Month"] == date.today().month]["Amount Usd"].sum()
            ]
        }
        df_summary = pd.DataFrame(summary)
        st.dataframe(df_summary.style.set_properties(**{
            'font-weight': 'bold', 'background-color': '#fce4ec', 'color': '#000'
        }), use_container_width=True)

        st.divider()
        st.markdown("### 📈 Weekly NGN Expenses")
        chart = df_filtered.groupby("Week")["Amount Ngn"].sum().reset_index()
        fig = px.bar(chart, x="Week", y="Amount Ngn", title="Weekly NGN Expenses",
                     color_discrete_sequence=px.colors.sequential.Reds)
        st.plotly_chart(fig, use_container_width=True)

        # --- Export ---
        st.markdown("### 📤 Export Summary")
        excel_out = io.BytesIO()
        with pd.ExcelWriter(excel_out, engine="xlsxwriter") as writer:
            df_summary.to_excel(writer, index=False, sheet_name="Summary")
            chart.to_excel(writer, index=False, sheet_name="Chart")

        st.download_button("⬇️ Download Excel", data=excel_out.getvalue(),
                           file_name="expense_summary.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        if st.button("📄 Download PDF", key="pdf_expense"):
            pdf = FPDF()
            pdf.add_page()
            pdf.image("assets/salmnine_logo.png", x=10, y=8, w=30)
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "Salmnine Investment Ltd", ln=True, align="C")
            pdf.set_font("Arial", size=12)
            pdf.cell(0, 10, "Expense Summary Report", ln=True, align="C")
            pdf.ln(10)
            for i, row in df_summary.iterrows():
                pdf.set_font("Arial", style="B", size=11)
                pdf.cell(0, 10, f"{row['Period']}: NGN = {row['Amount NGN']:,.2f}, USD = ${row['Amount USD']:,.2f}", ln=True)
            pdf.set_y(-20)
            pdf.set_font("Arial", "I", 8)
            pdf.cell(0, 10, "Generated by Salmnine Daily Tracker", 0, 0, "C")
            pdf_out = io.BytesIO()
            pdf.output(pdf_out)
            st.download_button("📄 Download PDF", data=pdf_out.getvalue(),
                               file_name="expense_summary.pdf", mime="application/pdf")
    else:
        st.info("ℹ️ No data available in Expenses sheet.")