import streamlit as st
import pandas as pd
import plotly.express as px
from fpdf import FPDF
import io
import xlsxwriter
from openpyxl import load_workbook
from datetime import date

# --- Excel File Setup ---
EXCEL_FILE = "Daily Transaction.xlsx"
TAB_NAME = "usdngn"
DATABASE_FILE = "Database.xlsx"

# --- Cache Excel File Loading for Performance ---
@st.cache_data
def load_excel_file(file_path, sheet_name):
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except (FileNotFoundError, ValueError) as e:
        st.error(f"❌ Failed to load {sheet_name}: {e}")
        return pd.DataFrame()

# --- Load Dropdown Options ---
def load_dropdown_options():
    try:
        # Load Buying Clients
        df_clients = load_excel_file(DATABASE_FILE, "Client List")
        buying_clients = [""] + df_clients["Client Name"].dropna().unique().tolist() if not df_clients.empty else [""]

        # Load Selling Clients
        df_sellers = load_excel_file(DATABASE_FILE, "Seller List")
        selling_clients = [""] + df_sellers["Seller Name"].dropna().unique().tolist() if not df_sellers.empty else [""]

        # Load Transaction Types
        df_trans_types = load_excel_file(DATABASE_FILE, "Transaction Type")
        trans_types = [""] + df_trans_types["Transaction Type"].dropna().unique().tolist() if not df_trans_types.empty else [""]

        return buying_clients, selling_clients, trans_types
    except Exception as e:
        st.error(f"❌ Failed to load dropdown options: {e}")
        return [""], [""], [""]

# --- Define Form Defaults ---
FORM_DEFAULTS = {
    "trans_date": date.today(),
    "selling_client": "",
    "buy_rate": 0.0,
    "buy_amt": 0.0,
    "buy_amt_in_bank": 0.0,
    "var_buy_amt": 0.0,
    "buying_client": "",
    "transaction_type": "",
    "sell_rate": 0.0,
    "fcy_val": 0.0,
    "lcy_val": 0.0,
    "lcy_paid_to_client": 0.0,
    "date_paid": date.today(),
    "bank_paid_from": "",
    "fcy_paid_to_client": 0.0
}

def initialize_form_state():
    for key, value in FORM_DEFAULTS.items():
        if key not in st.session_state:
            st.session_state[key] = value

# Load dropdown options
buying_clients, selling_clients, trans_types = load_dropdown_options()

# --- Paginated Data Table ---
def paginate_dataframe(df, page_size, page_number):
    start_idx = (page_number - 1) * page_size
    end_idx = start_idx + page_size
    return df.iloc[start_idx:end_idx]

# --- Form UI ---
def run_page():
    st.subheader("📘 USD/NGN Transactions Overview")

    # Ensure session state is initialized on every run
    initialize_form_state()

    # --- Transaction Form ---
    st.markdown("### 📝 Add New Transaction")
    with st.form("usdngn_form", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            # Use a fallback value to avoid AttributeError
            trans_date_value = st.session_state.get("trans_date", FORM_DEFAULTS["trans_date"])
            trans_date = st.date_input("Date", value=trans_date_value, key="trans_date")
            
            selling_client_value = st.session_state.get("selling_client", FORM_DEFAULTS["selling_client"])
            selling_client = st.selectbox("Selling Client", options=selling_clients, 
                                        index=selling_clients.index(selling_client_value) if selling_client_value in selling_clients else 0, 
                                        key="selling_client")
            
            buy_rate_value = st.session_state.get("buy_rate", FORM_DEFAULTS["buy_rate"])
            buy_rate = st.number_input("Buy Rate", min_value=0.0, value=buy_rate_value, key="buy_rate")
            
            buy_amt_value = st.session_state.get("buy_amt", FORM_DEFAULTS["buy_amt"])
            buy_amt = st.number_input("Buy Amt", min_value=0.0, value=buy_amt_value, key="buy_amt")
            
            buy_amt_in_bank_value = st.session_state.get("buy_amt_in_bank", FORM_DEFAULTS["buy_amt_in_bank"])
            buy_amt_in_bank = st.number_input("Buy Amt in Bank", min_value=0.0, value=buy_amt_in_bank_value, key="buy_amt_in_bank")
            
            var_buy_amt_value = st.session_state.get("var_buy_amt", FORM_DEFAULTS["var_buy_amt"])
            var_buy_amt = st.number_input("Var Buy Amt", min_value=0.0, value=var_buy_amt_value, key="var_buy_amt")
            
            buying_client_value = st.session_state.get("buying_client", FORM_DEFAULTS["buying_client"])
            buying_client = st.selectbox("Buying Client", options=buying_clients, 
                                       index=buying_clients.index(buying_client_value) if buying_client_value in buying_clients else 0, 
                                       key="buying_client")
        with col2:
            transaction_type_value = st.session_state.get("transaction_type", FORM_DEFAULTS["transaction_type"])
            transaction_type = st.selectbox("Transaction Type", options=trans_types, 
                                          index=trans_types.index(transaction_type_value) if transaction_type_value in trans_types else 0, 
                                          key="transaction_type")
            
            sell_rate_value = st.session_state.get("sell_rate", FORM_DEFAULTS["sell_rate"])
            sell_rate = st.number_input("Sell Rate", min_value=0.0, value=sell_rate_value, key="sell_rate")
            
            fcy_val_value = st.session_state.get("fcy_val", FORM_DEFAULTS["fcy_val"])
            fcy_val = st.number_input("FCY Val", min_value=0.0, value=fcy_val_value, key="fcy_val")
            
            lcy_val_value = st.session_state.get("lcy_val", FORM_DEFAULTS["lcy_val"])
            lcy_val = st.number_input("LCY Val", min_value=0.0, value=lcy_val_value, key="lcy_val")
            
            lcy_paid_to_client_value = st.session_state.get("lcy_paid_to_client", FORM_DEFAULTS["lcy_paid_to_client"])
            lcy_paid_to_client = st.number_input("LCY Paid to Client", min_value=0.0, value=lcy_paid_to_client_value, key="lcy_paid_to_client")
        with col3:
            date_paid_value = st.session_state.get("date_paid", FORM_DEFAULTS["date_paid"])
            date_paid = st.date_input("Date Paid", value=date_paid_value, key="date_paid")
            
            bank_paid_from_value = st.session_state.get("bank_paid_from", FORM_DEFAULTS["bank_paid_from"])
            bank_paid_from = st.text_input("Bank Paid From", value=bank_paid_from_value, key="bank_paid_from")
            
            fcy_paid_to_client_value = st.session_state.get("fcy_paid_to_client", FORM_DEFAULTS["fcy_paid_to_client"])
            fcy_paid_to_client = st.number_input("FCY Paid to Client", min_value=0.0, value=fcy_paid_to_client_value, key="fcy_paid_to_client")

        # --- Display Automated Calculations ---
        st.markdown("#### 📊 Calculated Fields")
        calc_col1, calc_col2, calc_col3 = st.columns(3)
        
        # Perform calculations
        margin = st.session_state.sell_rate - st.session_state.buy_rate
        var_buy_amt_lcy_val = st.session_state.buy_amt - st.session_state.lcy_val
        fcy_outstanding = st.session_state.fcy_paid_to_client - st.session_state.fcy_val
        lcy_outstanding = st.session_state.lcy_paid_to_client - st.session_state.lcy_val
        profit = margin * st.session_state.fcy_val
        spread = 0.0
        if st.session_state.fcy_val != 0:
            spread = (st.session_state.lcy_paid_to_client / st.session_state.fcy_val) - st.session_state.sell_rate
        commission = st.session_state.fcy_val * spread

        with calc_col1:
            st.metric("Margin", f"{round(margin, 2)}")
            st.metric("Var Buy Amt - LCY Val", f"{round(var_buy_amt_lcy_val, 2)}")
            st.metric("FCY Outstanding", f"{round(fcy_outstanding, 2)}")
        with calc_col2:
            st.metric("LCY Outstanding", f"{round(lcy_outstanding, 2)}")
            st.metric("Profit", f"{round(profit, 2)}")
        with calc_col3:
            st.metric("Spread", f"{round(spread, 2)}")
            st.metric("Commission", f"{round(commission, 2)}")

        submitted = st.form_submit_button("✅ Submit Transaction")

    # --- Form Submission Logic ---
    if submitted:
        # Input validation
        if st.session_state.buy_rate < 0 or st.session_state.buy_amt < 0 or st.session_state.buy_amt_in_bank < 0 or \
           st.session_state.var_buy_amt < 0 or st.session_state.sell_rate < 0 or st.session_state.fcy_val < 0 or \
           st.session_state.lcy_val < 0 or st.session_state.lcy_paid_to_client < 0 or st.session_state.fcy_paid_to_client < 0:
            st.error("❌ All numeric fields must be non-negative.")
        elif not st.session_state.buying_client or not st.session_state.transaction_type:
            st.error("❌ Buying Client and Transaction Type are required.")
        else:
            try:
                new_row = {
                    "Date": st.session_state.trans_date.strftime("%Y-%m-%d"),
                    "Selling Client": st.session_state.selling_client,
                    "Buy Rate": round(st.session_state.buy_rate, 2),
                    "Buy Amt": round(st.session_state.buy_amt, 2),
                    "Buy amt in bank": round(st.session_state.buy_amt_in_bank, 2),
                    "Var Buy amt": round(st.session_state.var_buy_amt, 2),
                    "Buying Client": st.session_state.buying_client,
                    "Transaction type": st.session_state.transaction_type,
                    "Sell Rate": round(st.session_state.sell_rate, 2),
                    "Margin": round(margin, 2),
                    "FCY Val": round(st.session_state.fcy_val, 2),
                    "LCY Val": round(st.session_state.lcy_val, 2),
                    "Var buy amt -LCY Val": round(var_buy_amt_lcy_val, 2),
                    "LCY Paid to client": round(st.session_state.lcy_paid_to_client, 2),
                    "Date Paid": st.session_state.date_paid.strftime("%Y-%m-%d"),
                    "Bank Paid from": st.session_state.bank_paid_from,
                    "FCY Paid to client": round(st.session_state.fcy_paid_to_client, 2),
                    "FCY Outstanding": round(fcy_outstanding, 2),
                    "LCY Outstanding": round(lcy_outstanding, 2),
                    "Profit": round(profit, 2),
                    "Spread": round(spread, 2),
                    "Commission": round(commission, 2)
                }

                # Load existing data or create empty DataFrame
                try:
                    book = load_workbook(EXCEL_FILE)
                    if TAB_NAME in book.sheetnames:
                        df_existing = pd.read_excel(EXCEL_FILE, sheet_name=TAB_NAME)
                    else:
                        df_existing = pd.DataFrame(columns=new_row.keys())
                except (FileNotFoundError, ValueError):
                    df_existing = pd.DataFrame(columns=new_row.keys())

                # Append new row
                df_combined = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)

                # Write to Excel without overwriting other sheets
                try:
                    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        df_combined.to_excel(writer, sheet_name=TAB_NAME, index=False)
                    st.success("✅ Transaction submitted successfully!")
                except PermissionError:
                    st.error("❌ Cannot write to Excel file: File is open or lacks write permissions.")
                except Exception as e:
                    st.error(f"❌ Failed to write to Excel: {e}")

            except Exception as e:
                st.error(f"❌ Error processing submission: {e}")

    # --- Load Data ---
    df = load_excel_file(EXCEL_FILE, TAB_NAME)

    # Normalize headers
    if not df.empty:
        df.columns = [col.strip().title() for col in df.columns]

        # Convert Date column to datetime
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df[df["Date"].notna()]

        # Date range filter
        if not df.empty and df["Date"].notna().any():
            start = st.date_input("📅 Start Date", df["Date"].min().date(), key="start_usdngn")
            end = st.date_input("📅 End Date", df["Date"].max().date(), key="end_usdngn")
        else:
            start = st.date_input("📅 Start Date", date.today(), key="start_usdngn")
            end = st.date_input("📅 End Date", date.today(), key="end_usdngn")

        df_filtered = df[(df["Date"] >= pd.to_datetime(start)) & (df["Date"] <= pd.to_datetime(end))]

        # --- Paginated Data Table ---
        st.markdown("### 📊 Data Table")
        page_size = 10  # Number of rows per page
        total_rows = len(df_filtered)
        total_pages = (total_rows + page_size - 1) // page_size if total_rows > 0 else 1

        # Initialize page number in session state
        if "usdngn_page_number" not in st.session_state:
            st.session_state.usdngn_page_number = 1

        # Page navigation
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            if st.button("⬅️ Previous", key="usdngn_prev"):
                if st.session_state.usdngn_page_number > 1:
                    st.session_state.usdngn_page_number -= 1
        with col2:
            st.write(f"Page {st.session_state.usdngn_page_number} of {total_pages}")
        with col3:
            if st.button("Next ➡️", key="usdngn_next"):
                if st.session_state.usdngn_page_number < total_pages:
                    st.session_state.usdngn_page_number += 1

        # Display paginated data
        paginated_df = paginate_dataframe(df_filtered, page_size, st.session_state.usdngn_page_number)
        st.dataframe(paginated_df, use_container_width=True)

        # --- Summary Calculations ---
        st.markdown("### 📊 Transaction Summary")
        df_filtered["Profit"] = pd.to_numeric(df_filtered["Profit"], errors="coerce")
        df_filtered["Fcy Val"] = pd.to_numeric(df_filtered["Fcy Val"], errors="coerce")
        df_filtered["Week"] = df_filtered["Date"].dt.isocalendar().week
        df_filtered["Month"] = df_filtered["Date"].dt.month

        # Today's date for daily summary
        today = pd.to_datetime(date.today())

        summary = {
            "Period": ["Daily", "Weekly", "Monthly"],
            "Total Profit": [
                df_filtered[df_filtered["Date"].dt.date == today.date()]["Profit"].sum(),
                df_filtered[df_filtered["Week"] == today.isocalendar().week]["Profit"].sum(),
                df_filtered[df_filtered["Month"] == today.month]["Profit"].sum()
            ],
            "Total FCY Amount": [
                df_filtered[df_filtered["Date"].dt.date == today.date()]["Fcy Val"].sum(),
                df_filtered[df_filtered["Week"] == today.isocalendar().week]["Fcy Val"].sum(),
                df_filtered[df_filtered["Month"] == today.month]["Fcy Val"].sum()
            ]
        }
        df_summary = pd.DataFrame(summary)
        st.dataframe(df_summary.style.set_properties(**{
            'font-weight': 'bold', 'background-color': '#e3f2fd', 'color': '#000'
        }), use_container_width=True)

        # Chart: Profit over time
        if "Profit" in df_filtered.columns:
            st.markdown("### 📈 Profit Over Time")
            chart_data = df_filtered.groupby(df_filtered["Date"].dt.date)["Profit"].sum().reset_index()
            fig = px.bar(chart_data, x="Date", y="Profit", title="Total Profit Over Time",
                         color_discrete_sequence=px.colors.qualitative.Bold)
            st.plotly_chart(fig, use_container_width=True)

        # Export: Excel
        st.markdown("### 📤 Export Summary")
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df_filtered.to_excel(writer, index=False, sheet_name=TAB_NAME)
        st.download_button("⬇️ Download Excel", data=output.getvalue(),
                           file_name=TAB_NAME.replace(" ", "_").lower() + ".xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Export: PDF
        if st.button("⬇️ Download PDF", key="pdf_usdngn"):
            pdf = FPDF()
            pdf.add_page()
            pdf.image("assets/salmnine_logo.png", x=10, y=8, w=30)
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Salmnine Investment Ltd", ln=True, align="C")
            pdf.set_font("Arial", size=12)
            pdf.cell(0, 10, f"{TAB_NAME} Summary Report", ln=True, align="C")
            pdf.ln(10)

            if not df_filtered.empty:
                summary = df_filtered.describe().round(2)
                for col in summary.columns:
                    pdf.set_font("Arial", style="B", size=11)
                    pdf.cell(40, 10, txt=str(col), ln=True)
                    pdf.set_font("Arial", size=10)
                    for stat in summary.index:
                        val = summary.loc[stat, col]
                        pdf.cell(60, 8, f"{stat}: {val}", ln=True)
                    pdf.ln(5)
            else:
                pdf.cell(100, 10, txt="No data available.", ln=True)

            pdf.set_y(-20)
            pdf.set_font("Arial", "I", 8)
            pdf.cell(0, 10, "Generated by Salmnine Daily Transactions Suite", 0, 0, "C")
            pdf_output = io.BytesIO()
            pdf.output(pdf_output)
            st.download_button("📄 Download PDF", data=pdf_output.getvalue(),
                               file_name=TAB_NAME.replace(" ", "_").lower() + ".pdf", mime="application/pdf")
    else:
        st.info("ℹ️ No data available in USD/NGN Transactions sheet.")