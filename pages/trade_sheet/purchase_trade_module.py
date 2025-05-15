import streamlit as st
import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook
import plotly.express as px
from fpdf import FPDF
import io
import xlsxwriter

# --- Excel File Setup ---
TRADE_EXCEL_FILE = "Trade sheet.xlsx"
TRADE_TAB = "Purchase Trade"
DATABASE_FILE = "database.xlsx"

# --- Cache Excel File Loading for Performance ---
@st.cache_data
def load_excel_file(file_path, sheet_name):
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except (FileNotFoundError, ValueError) as e:
        st.error(f"❌ Failed to load {sheet_name}: {e}")
        return pd.DataFrame()

# --- Load Seller List from database.xlsx ---
df_sellers = load_excel_file(DATABASE_FILE, "Seller List")
seller_data = df_sellers.iloc[:, 0].dropna().tolist() if not df_sellers.empty else []

# --- Initialize Session State for Form Inputs ---
def initialize_form_state():
    defaults = {
        "purchase_date": date.today(),
        "purchase_trade_type": "Buy",
        "purchase_customer": "",
        "purchase_currency": "USD",
        "purchase_rate": 0.01,
        "purchase_ngn_amount": 0.0,
        "purchase_naira_paid": 0.0
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# Initialize session state
initialize_form_state()

# --- Form UI ---
st.subheader("🛒 Purchase Trade Entry")

st.write("Debug: Before form")  # Debugging to trace rendering
with st.form("purchase_form", clear_on_submit=False):
    st.write("Debug: Inside form")  # Debugging
    col1, col2 = st.columns(2)
    with col1:
        trade_date = st.date_input("Date", value=st.session_state.purchase_date, key="purchase_date")
        trade_type = st.selectbox("Buy/Sell", ["Buy", "Sell"], index=["Buy", "Sell"].index(st.session_state.purchase_trade_type) if st.session_state.purchase_trade_type in ["Buy", "Sell"] else 0, key="purchase_trade_type")
        customer = st.selectbox("Trade Customer", options=seller_data, index=seller_data.index(st.session_state.purchase_customer) if st.session_state.purchase_customer in seller_data else 0, key="purchase_customer")
        currency = st.text_input("Trade Currency", value=st.session_state.purchase_currency, key="purchase_currency")
        rate = st.number_input("Rate", min_value=0.01, step=0.01, value=st.session_state.purchase_rate, key="purchase_rate")
    with col2:
        ngn_amount = st.number_input("NGN Amount", min_value=0.0, step=100.0, value=st.session_state.purchase_ngn_amount, key="purchase_ngn_amount")
        trade_size = ngn_amount / rate if rate != 0 else 0
        st.markdown(f"**Trade Size (auto):** {trade_size:,.2f}")
        naira_paid = st.number_input("Naira Paid", min_value=0.0, step=100.0, value=st.session_state.purchase_naira_paid, key="purchase_naira_paid")
        naira_balance = ngn_amount - naira_paid
        st.markdown(f"**Naira Balance (auto):** {naira_balance:,.2f}")

    submitted = st.form_submit_button("✅ Submit Trade")

st.write("Debug: After form")  # Debugging

# --- Form Submission Logic ---
if submitted:
    # Input validation
    if st.session_state.purchase_rate <= 0:
        st.error("❌ Rate must be greater than 0.")
    else:
        try:
            new_row = {
                "date": st.session_state.purchase_date.strftime("%Y-%m-%d"),
                "buy/sell": st.session_state.purchase_trade_type,
                "trade customers": st.session_state.purchase_customer,
                "trade currency": st.session_state.purchase_currency,
                "trade size": round(trade_size, 2),
                "rate": round(st.session_state.purchase_rate, 2),
                "ngn amount": round(st.session_state.purchase_ngn_amount, 2),
                "naira paid": round(st.session_state.purchase_naira_paid, 2),
                "naira balance": round(naira_balance, 2)
            }

            # Load existing data or create empty DataFrame
            try:
                book = load_workbook(TRADE_EXCEL_FILE)
                if TRADE_TAB in book.sheetnames:
                    df_existing = pd.read_excel(TRADE_EXCEL_FILE, sheet_name=TRADE_TAB)
                else:
                    df_existing = pd.DataFrame(columns=new_row.keys())
            except (FileNotFoundError, ValueError):
                df_existing = pd.DataFrame(columns=new_row.keys())

            # Append new row
            df_combined = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)

            # Write to Excel without overwriting other sheets
            try:
                with pd.ExcelWriter(TRADE_EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df_combined.to_excel(writer, sheet_name=TRADE_TAB, index=False)
                st.success("✅ Trade entry submitted successfully!")
            except PermissionError:
                st.error("❌ Cannot write to Excel file: File is open or lacks write permissions.")
            except Exception as e:
                st.error(f"❌ Failed to write to Excel: {e}")

        except Exception as e:
            st.error(f"❌ Error processing submission: {e}")

# --- Fetch Data ---
try:
    df = pd.read_excel(TRADE_EXCEL_FILE, sheet_name=TRADE_TAB)
except (FileNotFoundError, ValueError):
    df = pd.DataFrame()

# ✅ Normalize headers
if not df.empty:
    df.columns = [col.strip().lower() for col in df.columns]

    # --- Filter ---
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].notna()]
    if not df.empty and df["date"].notna().any():
        start_date = st.date_input("📆 Filter From", df["date"].min().date())
        end_date = st.date_input("📆 Filter To", df["date"].max().date())
    else:
        start_date = st.date_input("📆 Filter From", date.today())
        end_date = st.date_input("📆 Filter To", date.today())
    
    df_filtered = df[(df["date"] >= pd.to_datetime(start_date)) & (df["date"] <= pd.to_datetime(end_date))]

    st.markdown("### 📊 Summary Table")
    df_filtered["Week"] = df_filtered["date"].dt.isocalendar().week
    df_filtered["Month"] = df_filtered["date"].dt.month

    summary = {
        "Period": ["Daily", "Weekly", "Monthly"],
        "Trade Size": [
            df_filtered[df_filtered["date"] == pd.to_datetime(date.today())]["trade size"].sum(),
            df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["trade size"].sum(),
            df_filtered[df_filtered["Month"] == date.today().month]["trade size"].sum()
        ],
        "Ngn Amount": [
            df_filtered[df_filtered["date"] == pd.to_datetime(date.today())]["ngn amount"].sum(),
            df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["ngn amount"].sum(),
            df_filtered[df_filtered["Month"] == date.today().month]["ngn amount"].sum()
        ]
    }
    df_summary = pd.DataFrame(summary)
    st.dataframe(df_summary.style.set_properties(**{
        'font-weight': 'bold',
        'background-color': '#e8f0fe',
        'color': '#000'
    }), use_container_width=True)

    st.divider()

    st.markdown("### 📈 Weekly Trade Size Summary")
    weekly_chart = df_filtered.groupby("Week")["trade size"].sum().reset_index()
    fig = px.bar(weekly_chart, x="Week", y="trade size", title="Filtered Weekly Trade Size",
                 color_discrete_sequence=px.colors.sequential.Teal)
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("### 📤 Export Summary & Chart")

    # --- Export to Excel ---
    excel_out = io.BytesIO()
    with pd.ExcelWriter(excel_out, engine="xlsxwriter") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Summary")
        weekly_chart.to_excel(writer, index=False, sheet_name="Chart")

    st.download_button("⬇️ Download Excel", data=excel_out.getvalue(),
                       file_name="purchase_trade_summary.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # --- Export to PDF ---
    if st.button("📄 Download PDF", key="pdf_purchase"):
        pdf = FPDF()
        pdf.add_page()
        pdf.image("assets/salmnine_logo.png", x=10, y=8, w=30)
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Salmnine Investment Ltd", ln=True, align="C")
        pdf.set_font("Arial", size=12)
        pdf.cell(0, 10, "Purchase Trade Summary Report", ln=True, align="C")
        pdf.ln(10)

        for i, row in df_summary.iterrows():
            pdf.set_font("Arial", style="B", size=11)
            pdf.cell(0, 10, f"{row['Period']}: Trade Size = ₦{row['Trade Size']:,.2f}, NGN Amount = ₦{row['Ngn Amount']:,.2f}", ln=True)
        pdf.set_y(-20)
        pdf.set_font("Arial", "I", 8)
        pdf.cell(0, 10, "Generated by Salmnine Trade Reporting Suite", 0, 0, "C")

        pdf_output = io.BytesIO()
        pdf.output(pdf_output)
        st.download_button("📄 Download PDF", data=pdf_output.getvalue(),
                           file_name="purchase_trade_summary.pdf", mime="application/pdf")
        st.success("PDF generated successfully!")
else:
    st.info("ℹ️ No data available in Purchase Trade sheet.")