import streamlit as st
import pandas as pd
from datetime import date
from openpyxl import load_workbook
import plotly.express as px
from fpdf import FPDF
import io
import xlsxwriter

# --- Excel File Setup ---
TRADE_EXCEL_FILE = "Trade sheet.xlsx"
TRADE_TAB = "USD Trade"
DATABASE_FILE = "database.xlsx"

# --- Cache Excel File Loading for Performance ---
@st.cache_data
def load_excel_file(file_path, sheet_name):
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except (FileNotFoundError, ValueError) as e:
        st.error(f"❌ Failed to load {sheet_name}: {e}")
        return pd.DataFrame()

# --- Load Client List from database.xlsx ---
df_clients = load_excel_file(DATABASE_FILE, "Client List")
client_list = df_clients.iloc[:, 0].dropna().tolist() if not df_clients.empty else []

# --- Initialize Session State for Form Inputs ---
def initialize_form_state():
    defaults = {
        "usd_date": date.today(),
        "usd_side": "Buy",
        "usd_customer": "",
        "usd_currency": "USD",
        "usd_trade_size": 0.0,
        "usd_sell_rate": 0.0,
        "usd_buy_rate": 0.0,
        "usd_received": 0.0,
        "usd_paid_out": 0.0,
        "usd_commission": 0.0
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# Initialize session state
initialize_form_state()

# --- Form UI ---
st.subheader("💵 USD Trade Entry")

st.write("Debug: Before form")  # Debugging to trace rendering
with st.form("usd_trade_form", clear_on_submit=False):
    st.write("Debug: Inside form")  # Debugging
    col1, col2 = st.columns(2)
    with col1:
        trade_date = st.date_input("Date", value=st.session_state.usd_date, key="usd_date")
        side = st.selectbox("Buy/Sell", ["Buy", "Sell"], index=["Buy", "Sell"].index(st.session_state.usd_side) if st.session_state.usd_side in ["Buy", "Sell"] else 0, key="usd_side")
        customer = st.selectbox("Trade Customer", client_list, index=client_list.index(st.session_state.usd_customer) if st.session_state.usd_customer in client_list else 0, key="usd_customer")
        currency = st.text_input("Trade Currency", value=st.session_state.usd_currency, key="usd_currency")
        trade_size = st.number_input("Trade Size", min_value=0.0, value=st.session_state.usd_trade_size, key="usd_trade_size")
        sell_rate = st.number_input("Sell Rate", min_value=0.0, value=st.session_state.usd_sell_rate, key="usd_sell_rate")
        buy_rate = st.number_input("Buy Rate", min_value=0.0, value=st.session_state.usd_buy_rate, key="usd_buy_rate")
    with col2:
        amount = trade_size * sell_rate
        st.markdown(f"**Amount (auto):** ₦{amount:,.2f}")
        usd_received = st.number_input("USD Received", min_value=0.0, value=st.session_state.usd_received, key="usd_received")
        usd_paid_out = st.number_input("USD Paid Out", min_value=0.0, value=st.session_state.usd_paid_out, key="usd_paid_out")
        commission = st.number_input("Commission", min_value=0.0, value=st.session_state.usd_commission, key="usd_commission")
        ngn_income = (sell_rate - buy_rate) * trade_size
        st.markdown(f"**NGN Income (auto):** ₦{ngn_income:,.2f}")

    submitted = st.form_submit_button("✅ Submit USD Trade")

st.write("Debug: After form")  # Debugging

# --- Form Submission Logic ---
if submitted:
    # Input validation
    if st.session_state.usd_trade_size <= 0 or st.session_state.usd_sell_rate <= 0 or st.session_state.usd_buy_rate <= 0:
        st.error("❌ Trade Size, Sell Rate, and Buy Rate must be greater than 0.")
    else:
        try:
            new_row = {
                "date": st.session_state.usd_date.strftime("%Y-%m-%d"),
                "buy/sell": st.session_state.usd_side,
                "trade customers": st.session_state.usd_customer,
                "trade currency": st.session_state.usd_currency,
                "trade size": round(st.session_state.usd_trade_size, 2),
                "sell rate": round(st.session_state.usd_sell_rate, 2),
                "amount": round(amount, 2),
                "usd received": round(st.session_state.usd_received, 2),
                "usd paid out": round(st.session_state.usd_paid_out, 2),
                "commission ngn": round(st.session_state.usd_commission, 2),
                "Ngn Income": round(ngn_income, 2),
                "buy rate": round(st.session_state.usd_buy_rate, 2)
            }

            # Load existing data or create empty DataFrame
            try:
                book = load_workbook(TRADE_EXCEL_FILE)
                if TRADE_TAB in book.sheetnames:
                    df_existing = pd.read_excel(TRADE_EXCEL_FILE, sheet_name=TRADE_TAB)
                else:
                    df_existing = pd.DataFrame(columns=new_row.keys())
            except (FileNotFoundError, ValueError):
                df_existing = pd.DataFrame(columns=new_row.keys())

            # Append new row
            df_combined = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)

            # Write to Excel without overwriting other sheets
            try:
                with pd.ExcelWriter(TRADE_EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df_combined.to_excel(writer, sheet_name=TRADE_TAB, index=False)
                st.success("✅ Trade submitted successfully!")
            except PermissionError:
                st.error("❌ Cannot write to Excel file: File is open or lacks write permissions.")
            except Exception as e:
                st.error(f"❌ Failed to write to Excel: {e}")

        except Exception as e:
            st.error(f"❌ Error processing submission: {e}")

# --- Load and Normalize Data ---
try:
    df = pd.read_excel(TRADE_EXCEL_FILE, sheet_name=TRADE_TAB)
except (FileNotFoundError, ValueError):
    df = pd.DataFrame()

# Normalize headers
if not df.empty:
    df.columns = [col.strip().lower() for col in df.columns]

    # --- Summary & Filter ---
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df[df["date"].notna()]
    if not df.empty and df["date"].notna().any():
        start = st.date_input("📅 Start Date", df["date"].min().date())
        end = st.date_input("📅 End Date", df["date"].max().date())
    else:
        start = st.date_input("📅 Start Date", date.today())
        end = st.date_input("📅 End Date", date.today())
    
    df_filtered = df[(df["date"] >= pd.to_datetime(start)) & (df["date"] <= pd.to_datetime(end))]

    # ✅ Convert numerics
    for col in ["trade size", "amount", "ngn income"]:
        df_filtered[col] = pd.to_numeric(df_filtered[col], errors="coerce")

    st.markdown("### 📊 Summary")
    df_filtered["Week"] = df_filtered["date"].dt.isocalendar().week
    df_filtered["Month"] = df_filtered["date"].dt.month

    summary = {
        "Period": ["Daily", "Weekly", "Monthly"],
        "Trade Size": [
            df_filtered[df_filtered["date"] == pd.to_datetime(date.today())]["trade size"].sum(),
            df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["trade size"].sum(),
            df_filtered[df_filtered["Month"] == date.today().month]["trade size"].sum()
        ],
        "Amount": [
            df_filtered[df_filtered["date"] == pd.to_datetime(date.today())]["amount"].sum(),
            df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["amount"].sum(),
            df_filtered[df_filtered["Month"] == date.today().month]["amount"].sum()
        ],
        "Ngn Income": [
            df_filtered[df_filtered["date"] == pd.to_datetime(date.today())]["ngn income"].sum(),
            df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["ngn income"].sum(),
            df_filtered[df_filtered["Month"] == date.today().month]["ngn income"].sum()
        ]
    }
    df_summary = pd.DataFrame(summary)
    st.dataframe(df_summary.style.set_properties(**{
        'font-weight': 'bold', 'background-color': '#f0f4c3', 'color': '#000'
    }), use_container_width=True)

    st.divider()
    st.markdown("### 📈 Weekly Trade Size")
    chart = df_filtered.groupby("Week")["trade size"].sum().reset_index()
    fig = px.bar(chart, x="Week", y="trade size", title="Weekly Trade Volume",
                 color_discrete_sequence=px.colors.qualitative.Safe)
    st.plotly_chart(fig, use_container_width=True)

    # --- Export ---
    st.markdown("### 📤 Export Summary")
    excel_out = io.BytesIO()
    with pd.ExcelWriter(excel_out, engine="xlsxwriter") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Summary")
        chart.to_excel(writer, index=False, sheet_name="Chart")

    st.download_button("⬇️ Download Excel", data=excel_out.getvalue(),
                       file_name="usd_trade_summary.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if st.button("📄 Download PDF", key="pdf_usd"):
        pdf = FPDF()
        pdf.add_page()
        pdf.image("assets/salmnine_logo.png", x=10, y=8, w=30)
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Salmnine Investment Ltd", ln=True, align="C")
        pdf.set_font("Arial", size=12)
        pdf.cell(0, 10, "USD Trade Summary Report", ln=True, align="C")
        pdf.ln(10)
        for i, row in df_summary.iterrows():
            pdf.set_font("Arial", style="B", size=11)
            pdf.cell(0, 10, f"{row['Period']}: Trade Size = ₦{row['Trade Size']:,.2f}, "
                            f"Amount = ₦{row['Amount']:,.2f}, NGN Income = ₦{row['Ngn Income']:,.2f}", ln=True)
        pdf.set_y(-20)
        pdf.set_font("Arial", "I", 8)
        pdf.cell(0, 10, "Generated by Salmnine Trade Suite", 0, 0, "C")
        pdf_out = io.BytesIO()
        pdf.output(pdf_out)
        st.download_button("📄 Download PDF", data=pdf_out.getvalue(),
                           file_name="usd_trade_summary.pdf", mime="application/pdf")
        st.success("PDF generated successfully!")
else:
    st.warning("No trades found for the selected date range.")

# --- Logout ---
if st.button("🚪 Logout", key="logout_usd"):
    st.session_state.auth["trade_sheet"] = {"authenticated": False}
    st.success("Logged out successfully")
    st.experimental_rerun()