def render_swap_trade_trade():
    import streamlit as st
    import pandas as pd
    from datetime import date
    from openpyxl import load_workbook
    import plotly.express as px
    from fpdf import FPDF
    import io
    import xlsxwriter
    
    # --- Excel File Setup ---
    TRADE_EXCEL_FILE = "Trade sheet.xlsx"
    TRADE_TAB = "Swap Trade"
    DATABASE_FILE = "database.xlsx"
    
    # --- Cache Excel File Loading for Performance ---
    @st.cache_data
    def load_excel_file(file_path, sheet_name):
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        except (FileNotFoundError, ValueError) as e:
            st.error(f"❌ Failed to load {sheet_name}: {e}")
            return pd.DataFrame()
    
    # --- Load Client List from database.xlsx ---
    df_clients = load_excel_file(DATABASE_FILE, "Client List")
    client_list = df_clients.iloc[:, 0].dropna().tolist() if not df_clients.empty else []
    
    # --- Initialize Session State for Form Inputs ---
    def initialize_form_state():
        defaults = {
            "swap_date": date.today(),
            "swap_client": "",
            "swap_usd_received": 0.0,
            "swap_charges_pct": 0.0,
            "swap_usdt_paid": 0.0,
            "swap_usd_status": "Pending",
            "swap_usdt_status": "Pending",
            "swap_date_received": date.today(),
            "swap_date_sent": date.today()
        }
        for key, value in defaults.items():
            if key not in st.session_state:
                st.session_state[key] = value
    
    # Initialize session state
    initialize_form_state()
    
    # --- Form UI ---
    st.subheader("🔄 Swap Trade Entry")
    
    st.write("Debug: Before form")  # Debugging to trace rendering
    with st.form("swap_trade_form", clear_on_submit=False):
        st.write("Debug: Inside form")  # Debugging
        col1, col2 = st.columns(2)
        with col1:
            trade_date = st.date_input("Date", value=st.session_state.swap_date, key="swap_date")
            client_name = st.selectbox("Client", client_list, index=client_list.index(st.session_state.swap_client) if st.session_state.swap_client in client_list else 0, key="swap_client")
            usd_received = st.number_input("USD Received", min_value=0.0, value=st.session_state.swap_usd_received, key="swap_usd_received")
            charges_pct = st.number_input("Charges (%)", min_value=0.0, step=0.1, value=st.session_state.swap_charges_pct, key="swap_charges_pct")
            charges_usdt = (charges_pct / 100) * usd_received
            usdt_due = usd_received - charges_usdt
            st.markdown(f"**Charges (USDT):** {charges_usdt:,.2f}")
            st.markdown(f"**USDT Due:** {usdt_due:,.2f}")
        with col2:
            usdt_paid = st.number_input("USDT Paid", min_value=0.0, value=st.session_state.swap_usdt_paid, key="swap_usdt_paid")
            usd_status = st.selectbox("USD Status", ["Pending", "Completed"], index=["Pending", "Completed"].index(st.session_state.swap_usd_status) if st.session_state.swap_usd_status in ["Pending", "Completed"] else 0, key="swap_usd_status")
            usdt_status = st.selectbox("USDT Status", ["Pending", "Completed"], index=["Pending", "Completed"].index(st.session_state.swap_usdt_status) if st.session_state.swap_usdt_status in ["Pending", "Completed"] else 0, key="swap_usdt_status")
            date_received = st.date_input("Date Received", value=st.session_state.swap_date_received, key="swap_date_received")
            date_sent = st.date_input("Date Sent", value=st.session_state.swap_date_sent, key="swap_date_sent")
            net_profit = usd_received - usdt_due
            st.markdown(f"**Net Profit (auto):** {net_profit:,.2f}")
    
        submitted = st.form_submit_button("✅ Submit Swap Trade")
    
    st.write("Debug: After form")  # Debugging
    
    # --- Form Submission Logic ---
    if submitted:
        # Input validation
        if st.session_state.swap_usd_received < 0 or st.session_state.swap_charges_pct < 0 or st.session_state.swap_usdt_paid < 0:
            st.error("❌ USD Received, Charges %, and USDT Paid must be non-negative.")
        else:
            try:
                new_row = {
                    "Date": st.session_state.swap_date.strftime("%Y-%m-%d"),
                    "Client": st.session_state.swap_client,
                    "USD Received": round(st.session_state.swap_usd_received, 2),
                    "Charges %": round(st.session_state.swap_charges_pct, 2),
                    "Charges (USDT)": round(charges_usdt, 2),
                    "USDT Due": round(usdt_due, 2),
                    "USDT Paid": round(st.session_state.swap_usdt_paid, 2),
                    "USD Status": st.session_state.swap_usd_status,
                    "USDT Status": st.session_state.swap_usdt_status,
                    "Date Received": st.session_state.swap_date_received.strftime("%Y-%m-%d"),
                    "Date Sent": st.session_state.swap_date_sent.strftime("%Y-%m-%d"),
                    "Net Profit": round(net_profit, 2)
                }
    
                # Load existing data or create empty DataFrame
                try:
                    book = load_workbook(TRADE_EXCEL_FILE)
                    if TRADE_TAB in book.sheetnames:
                        df_existing = pd.read_excel(TRADE_EXCEL_FILE, sheet_name=TRADE_TAB)
                    else:
                        df_existing = pd.DataFrame(columns=new_row.keys())
                except (FileNotFoundError, ValueError):
                    df_existing = pd.DataFrame(columns=new_row.keys())
    
                # Append new row
                df_combined = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)
    
                # Write to Excel without overwriting other sheets
                try:
                    with pd.ExcelWriter(TRADE_EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        df_combined.to_excel(writer, sheet_name=TRADE_TAB, index=False)
                    st.success("✅ Swap trade submitted successfully!")
                except PermissionError:
                    st.error("❌ Cannot write to Excel file: File is open or lacks write permissions.")
                except Exception as e:
                    st.error(f"❌ Failed to write to Excel: {e}")
    
            except Exception as e:
                st.error(f"❌ Error processing submission: {e}")
    
    # --- Load Data ---
    try:
        df = pd.read_excel(TRADE_EXCEL_FILE, sheet_name=TRADE_TAB)
    except (FileNotFoundError, ValueError):
        df = pd.DataFrame()
    
    # Normalize headers
    if not df.empty:
        df.columns = [col.strip().title() for col in df.columns]
    
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df[df["Date"].notna()]
        if not df.empty and df["Date"].notna().any():
            start = st.date_input("📅 Start Date", df["Date"].min().date())
            end = st.date_input("📅 End Date", df["Date"].max().date())
        else:
            start = st.date_input("📅 Start Date", date.today())
            end = st.date_input("📅 End Date", date.today())
        
        df_filtered = df[(df["Date"] >= pd.to_datetime(start)) & (df["Date"] <= pd.to_datetime(end))]
    
        for col in ["Net Profit", "Usdt Due"]:
            df_filtered[col] = pd.to_numeric(df_filtered[col], errors="coerce")
    
        st.markdown("### 📊 Summary")
        df_filtered["Week"] = df_filtered["Date"].dt.isocalendar().week
        df_filtered["Month"] = df_filtered["Date"].dt.month
    
        summary = {
            "Period": ["Daily", "Weekly", "Monthly"],
            "Net Profit": [
                df_filtered[df_filtered["Date"] == pd.to_datetime(date.today())]["Net Profit"].sum(),
                df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["Net Profit"].sum(),
                df_filtered[df_filtered["Month"] == date.today().month]["Net Profit"].sum()
            ],
            "USDT Due": [
                df_filtered[df_filtered["Date"] == pd.to_datetime(date.today())]["Usdt Due"].sum(),
                df_filtered[df_filtered["Week"] == date.today().isocalendar().week]["Usdt Due"].sum(),
                df_filtered[df_filtered["Month"] == date.today().month]["Usdt Due"].sum()
            ]
        }
        df_summary = pd.DataFrame(summary)
        st.dataframe(df_summary.style.set_properties(**{
            'font-weight': 'bold', 'background-color': '#fff3e0', 'color': '#000'
        }), use_container_width=True)
    
        st.divider()
        st.markdown("### 📈 Weekly Net Profit")
        chart = df_filtered.groupby("Week")["Net Profit"].sum().reset_index()
        fig = px.bar(chart, x="Week", y="Net Profit", title="Weekly Net Profit",
                     color_discrete_sequence=px.colors.sequential.Agsunset)
        st.plotly_chart(fig, use_container_width=True)
    
        st.markdown("### 📤 Export Summary")
        excel_out = io.BytesIO()
        with pd.ExcelWriter(excel_out, engine="xlsxwriter") as writer:
            df_summary.to_excel(writer, index=False, sheet_name="Summary")
            chart.to_excel(writer, index=False, sheet_name="Chart")
    
        st.download_button("⬇️ Download Excel", data=excel_out.getvalue(),
                           file_name="swap_trade_summary.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
        if st.button("📄 Download PDF", key="pdf_swap"):
            pdf = FPDF()
            pdf.add_page()
            pdf.image("assets/salmnine_logo.png", x=10, y=8, w=30)
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "Salmnine Investment Ltd", ln=True, align="C")
            pdf.set_font("Arial", size=12)
            pdf.cell(0, 10, "Swap Trade Summary Report", ln=True, align="C")
            pdf.ln(10)
            for i, row in df_summary.iterrows():
                pdf.set_font("Arial", style="B", size=11)
                pdf.cell(0, 10, f"{row['Period']}: Net Profit = ${row['Net Profit']:,.2f}, "
                                f"USDT Due = {row['Usdt Due']:,.2f}", ln=True)
            pdf.set_y(-20)
            pdf.set_font("Arial", "I", 8)
            pdf.cell(0, 10, "Generated by Salmnine Trade Suite", 0, 0, "C")
            pdf_out = io.BytesIO()
            pdf.output(pdf_out)
            st.download_button("📄 Download PDF", data=pdf_out.getvalue(),
                               file_name="swap_trade_summary.pdf", mime="application/pdf")
            st.success("PDF generated successfully!")
    else:
        st.info("ℹ️ No data available in Swap Trade sheet.")
