def render_ghs_trade():
    import streamlit as st
    import pandas as pd
    from datetime import date
    from openpyxl import load_workbook
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
    
    TRADE_TAB = "GHS Trade"
    EXCEL_FILE = "Trade sheet.xlsx"
    DATABASE_FILE = "database.xlsx"
    
    # --- Initialize Session State for Form Inputs ---
    def initialize_form_state():
        defaults = {
            "ghs_date": date.today(),
            "ghs_side": "Buy",
            "ghs_customer": "",
            "ghs_currency": "GHS",
            "ghs_trade_size": 0.0,
            "ghs_sell_rate": 0.01,
            "ghs_received": 0.0,
            "ghs_paid_out": 0.0,
            "ghs_commission": 0.0,
            "ghs_amount2": 0.0,
            "ghs_buy_rate": 0.01,
            "ghs_customer2": ""
        }
        for key, value in defaults.items():
            if key not in st.session_state:
                st.session_state[key] = value
    
    # Initialize session state
    initialize_form_state()
    
    # --- Cache Excel File Loading for Performance ---
    @st.cache_data
    def load_excel_file(file_path, sheet_name):
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        except (FileNotFoundError, ValueError) as e:
            st.error(f"❌ Failed to load {sheet_name}: {e}")
            return pd.DataFrame()
    
    # --- Load Client and Seller Lists ---
    df_clients = load_excel_file(DATABASE_FILE, "Client List")
    client_list = df_clients.iloc[:, 0].dropna().tolist() if not df_clients.empty else []
    
    df_sellers = load_excel_file(DATABASE_FILE, "Seller List")
    seller_list = df_sellers.iloc[:, 0].dropna().tolist() if not df_sellers.empty else []
    
    # --- UI Header ---
    st.markdown("""
    <div style='background-color:#E8F4FD;padding:12px;border-radius:8px;margin-bottom:15px'>
        <h3 style='color:#1B4F72'>🇬🇭 GHS Trade Entry</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # --- Form ---
    st.write("Debug: Before form")  # Debugging to trace rendering
    with st.form("ghs_trade_form", clear_on_submit=False):
        st.write("Debug: Inside form")  # Debugging
        col1, col2 = st.columns(2)
        with col1:
            trade_date = st.date_input("Date", value=st.session_state.ghs_date, key="ghs_date")
            side = st.selectbox("Buy/Sell", ["Buy", "Sell"], index=["Buy", "Sell"].index(st.session_state.ghs_side) if st.session_state.ghs_side in ["Buy", "Sell"] else 0, key="ghs_side")
            customer = st.selectbox("Trade Customer", client_list, index=client_list.index(st.session_state.ghs_customer) if st.session_state.ghs_customer in client_list else 0, key="ghs_customer")
            currency = st.text_input("Trade Currency", value=st.session_state.ghs_currency, key="ghs_currency")
            trade_size = st.number_input("Trade Size", min_value=0.0, value=st.session_state.ghs_trade_size, key="ghs_trade_size")
            sell_rate = st.number_input("Sell Rate", min_value=0.01, value=st.session_state.ghs_sell_rate, key="ghs_sell_rate")
            amount_ghs = trade_size * sell_rate
            st.markdown(f"**Amount GHS (auto):** ₵{amount_ghs:,.2f}")
            received = st.number_input("Received", min_value=0.0, value=st.session_state.ghs_received, key="ghs_received")
            paid_out = st.number_input("Paid Out", min_value=0.0, value=st.session_state.ghs_paid_out, key="ghs_paid_out")
            commission = st.number_input("Commission", min_value=0.0, value=st.session_state.ghs_commission, key="ghs_commission")
        with col2:
            amount_ghs2 = st.number_input("Amount GHS 2", min_value=0.0, value=st.session_state.ghs_amount2, key="ghs_amount2")
            buy_rate = st.number_input("Buy Rate", min_value=0.01, value=st.session_state.ghs_buy_rate, key="ghs_buy_rate")
            trade_size2 = amount_ghs2 / buy_rate if buy_rate != 0 else 0
            st.markdown(f"**Trade Size 2 (auto):** {trade_size2:,.2f}")
            income = trade_size2 - trade_size
            st.markdown(f"**Income (auto):** ₵{income:,.2f}")
            customer2 = st.selectbox("Trade Customer 2", seller_list, index=seller_list.index(st.session_state.ghs_customer2) if st.session_state.ghs_customer2 in seller_list else 0, key="ghs_customer2")
    
        submitted = st.form_submit_button("✅ Submit GHS Trade")
    
    st.write("Debug: After form")  # Debugging
    
    # --- Form Submission Logic ---
    if submitted:
        # Input validation
        if st.session_state.ghs_trade_size <= 0 or st.session_state.ghs_sell_rate <= 0 or st.session_state.ghs_buy_rate <= 0:
            st.error("❌ Trade Size, Sell Rate, and Buy Rate must be greater than 0.")
        else:
            try:
                new_row = {
                    "Date": st.session_state.ghs_date.strftime("%Y-%m-%d"),
                    "Buy/Sell": st.session_state.ghs_side,
                    "Trade Customer": st.session_state.ghs_customer,
                    "Trade Currency": st.session_state.ghs_currency,
                    "Trade Size": round(st.session_state.ghs_trade_size, 2),
                    "Sell Rate": round(st.session_state.ghs_sell_rate, 2),
                    "Amount GHS": round(amount_ghs, 2),
                    "Received": round(st.session_state.ghs_received, 2),
                    "Paid Out": round(st.session_state.ghs_paid_out, 2),
                    "Commission": round(st.session_state.ghs_commission, 2),
                    "Income": round(income, 2),
                    "Buy Rate": round(st.session_state.ghs_buy_rate, 2),
                    "Amount GHS 2": round(st.session_state.ghs_amount2, 2),
                    "Trade Customer 2": st.session_state.ghs_customer2,
                    "Trade Size 2": round(trade_size2, 2)
                }
    
                # Load existing data or create empty DataFrame
                try:
                    book = load_workbook(EXCEL_FILE)
                    if TRADE_TAB in book.sheetnames:
                        df_existing = pd.read_excel(EXCEL_FILE, sheet_name=TRADE_TAB)
                    else:
                        df_existing = pd.DataFrame(columns=new_row.keys())
                except (FileNotFoundError, ValueError):
                    df_existing = pd.DataFrame(columns=new_row.keys())
    
                # Append new row
                df_combined = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)
    
                # Write to Excel without overwriting other sheets
                try:
                    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        df_combined.to_excel(writer, sheet_name=TRADE_TAB, index=False)
                    st.success("✅ GHS trade submitted successfully!")
                except PermissionError:
                    st.error("❌ Cannot write to Excel file: File is open or lacks write permissions.")
                except Exception as e:
                    st.error(f"❌ Failed to write to Excel: {e}")
    
            except Exception as e:
                st.error(f"❌ Error processing submission: {e}")
    
    # --- Editable Table ---
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=TRADE_TAB)
        df.columns = [col.strip().title().replace("  ", " ") for col in df.columns]
        df.rename(columns={"Amount Ghs": "Amount GHS", "Amount Ghs 2": "Amount GHS 2"}, inplace=True)
    except (FileNotFoundError, ValueError):
        df = pd.DataFrame()
    
    if not df.empty:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df[df["Date"].notna()]
        
        # Robust date range selection
        if not df.empty and df["Date"].notna().any():
            start = st.date_input("📅 Start Date", df["Date"].min().date())
            end = st.date_input("📅 End Date", df["Date"].max().date())
        else:
            start = st.date_input("📅 Start Date", date.today())
            end = st.date_input("📅 End Date", date.today())
        
        df_filtered = df[(df["Date"] >= pd.to_datetime(start)) & (df["Date"] <= pd.to_datetime(end))]
    
        st.markdown("### 📋 Editable Table")
        gb = GridOptionsBuilder.from_dataframe(df_filtered)
        gb.configure_pagination()
        gb.configure_default_column(editable=True, filter=True, resizable=True)
        grid_response = AgGrid(
            df_filtered,
            gridOptions=gb.build(),
            update_mode=GridUpdateMode.VALUE_CHANGED,
            fit_columns_on_grid_load=True,
            height=350
        )
    
        updated_df = grid_response["data"]
        if not updated_df.equals(df_filtered):
            try:
                with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    updated_df.to_excel(writer, sheet_name=TRADE_TAB, index=False)
                st.success("✅ Table updates saved to Excel!")
            except PermissionError:
                st.error("❌ Cannot save table updates: File is open or lacks write permissions.")
            except Exception as e:
                st.error(f"❌ Failed to save table updates: {e}")
    else:
        st.info("ℹ️ No data available in GHS Trade sheet.")
